#!/usr/bin/env python3
"""
runZero Vulnerability Report by Asset Type
-------------------------------------------
Connects to the runZero API, pulls all medium/high/critical vulnerabilities,
and generates a formatted Excel workbook with one tab per asset-type group.

Output file is written to the current
working directory with the current date in the filename.

Dependencies:
    pip install requests openpyxl
"""

import json
import os
import re
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Dict, List, Tuple

import requests
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# CONFIGURATION — paste your org API key below
# ──────────────────────────────────────────────

RUNZERO_BASE_URL = "https://console.runzero.com/api/v1.0"
RUNZERO_ORG_TOKEN = "PASTE_YOUR_ORG_API_KEY_HERE"  # <-- PASTE YOUR ORG API KEY HERE

# ──────────────────────────────────────────────
# ASSET TYPE QUERIES
# ──────────────────────────────────────────────

SEVERITY_FILTER = "(severity:critical OR severity:high OR severity:medium)"

TYPE_QUERIES: Dict[str, str] = {
    "Endpoints": f'{SEVERITY_FILTER} AND (type:"Laptop" OR type:"Desktop" OR type:"Workstation" OR type:"Tablet")',
    "Hypervisors": f'{SEVERITY_FILTER} AND type:"Hypervisor"',
    "Virtual Machines": f'{SEVERITY_FILTER} AND type:"Virtual Machine"',
    "Servers": f'{SEVERITY_FILTER} AND type:"Server"',
    "Network Appliances": f'{SEVERITY_FILTER} AND (type:"Network" OR type:"Firewall" OR type:"Switch" OR type:"Router" OR type:"Load Balancer" OR type:"Wireless AP" OR type:"Access Point")',
    "Storage - NAS": f'{SEVERITY_FILTER} AND (type:"NAS" OR type:"Storage" OR type:"SAN")',
    "Printers": f'{SEVERITY_FILTER} AND type:"Printer"',
    "Phones - VoIP": f'{SEVERITY_FILTER} AND (type:"Phone" OR type:"VoIP")',
    "IoT - OT": f'{SEVERITY_FILTER} AND (type:"Camera" OR type:"IoT" OR type:"SCADA")',
    "BMC - IPMI": f'{SEVERITY_FILTER} AND (type:"BMC" OR type:"IPMI")',
}

# ──────────────────────────────────────────────
# STYLING CONSTANTS
# ──────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

SEVERITY_FILLS = {
    "critical": PatternFill(start_color="FF4D4D", end_color="FF4D4D", fill_type="solid"),
    "high": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "medium": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
}
SEVERITY_FONTS = {
    "critical": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
    "high": Font(name="Calibri", size=10, bold=True, color="000000"),
    "medium": Font(name="Calibri", size=10, color="000000"),
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMNS = [
    "Vulnerability Name",
    "Criticality",
    "CVSS Score",
    "# Affected Hosts",
    "Affected Hosts",
    "Description",
    "Remediation Steps",
]

COL_WIDTHS = [45, 12, 12, 16, 55, 60, 60]


# ──────────────────────────────────────────────
# API FUNCTIONS
# ──────────────────────────────────────────────

def _stream_group(group_name: str, search_query: str, tmp_dir: str) -> Tuple[str, str, int]:
    """
    Stream JSONL for one type-group into a temp file on disk.
    Retries on connection drop until the full stream completes.
    """
    url = f"{RUNZERO_BASE_URL}/export/org/vulnerabilities.jsonl"
    headers = {
        "Authorization": f"Bearer {RUNZERO_ORG_TOKEN}",
        "Accept": "application/json",
        "User-Agent": "runzero-vuln-report/1.0",
    }
    params = {"search": search_query}
    safe_name = re.sub(r'[^\w\-]', '_', group_name)
    filepath = os.path.join(tmp_dir, f"{safe_name}.jsonl")

    max_retries = 10
    for attempt in range(1, max_retries + 1):
        if attempt > 1:
            # Exponential backoff: 10s, 20s, 40s, 80s... capped at 120s
            wait = min(10 * (2 ** (attempt - 2)), 120)
            print(f"  [{group_name}] Waiting {wait}s before retry...", flush=True)
            time.sleep(wait)

        try:
            resp = requests.get(url, headers=headers, params=params, stream=True, timeout=(60, 1800))
            resp.raise_for_status()
        except requests.RequestException as e:
            print(f"  [{group_name}] Connection failed (attempt {attempt}): {e}", flush=True)
            if attempt < max_retries:
                continue
            return (group_name, filepath, 0)

        count = 0
        try:
            with open(filepath, "w") as f:
                for line in resp.iter_lines(decode_unicode=True):
                    if not line:
                        continue
                    f.write(line + "\n")
                    count += 1
                    if count % 5000 == 0:
                        print(f"  [{group_name}] ... {count} records", flush=True)
            # Stream completed successfully
            print(f"  [{group_name}] Done — {count} records.", flush=True)
            return (group_name, filepath, count)

        except (requests.exceptions.ChunkedEncodingError,
                requests.exceptions.ConnectionError) as e:
            if attempt < max_retries:
                print(f"  [{group_name}] Dropped at {count} records, retrying ({attempt}/{max_retries})...", flush=True)
            else:
                print(f"  [{group_name}] Dropped at {count} records after {max_retries} attempts. Using partial data.", flush=True)
                return (group_name, filepath, count)

    return (group_name, filepath, 0)


def fetch_vulnerabilities(tmp_dir: str) -> Dict[str, str]:
    """
    Fetch vulnerabilities for all type groups concurrently.
    Returns {group_name: filepath} for groups that have data.
    """
    print(f"Streaming all groups in parallel (max 4 concurrent)...", flush=True)

    group_files: Dict[str, str] = {}

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = {
            executor.submit(_stream_group, name, query, tmp_dir): name
            for name, query in TYPE_QUERIES.items()
        }

        for future in as_completed(futures):
            group_name, filepath, count = future.result()
            if count > 0 and os.path.exists(filepath):
                group_files[group_name] = filepath

    total = 0
    for fp in group_files.values():
        with open(fp, "r") as f:
            total += sum(1 for _ in f)
    print(f"\nTotal: {total} vulnerability records across {len(group_files)} groups.", flush=True)
    return group_files


# ──────────────────────────────────────────────
# DATA PROCESSING — deduplicate vulns, aggregate hosts
# ──────────────────────────────────────────────

def _extract_host(rec: dict) -> str:
    """Extract a host identifier (IP + hostname) from a vulnerability record."""
    addresses = rec.get("addresses") or []
    service_addr = rec.get("vulnerability_service_address", "") or ""
    if service_addr:
        ip = service_addr
    elif isinstance(addresses, list) and addresses:
        ip = str(addresses[0])
    elif isinstance(addresses, str) and addresses:
        ip = addresses.split()[0]
    else:
        ip = "N/A"

    names = rec.get("names") or []
    if isinstance(names, list) and names:
        hostname = str(names[0])
    elif isinstance(names, str) and names:
        hostname = names.split()[0]
    else:
        hostname = ""

    if hostname and hostname != ip:
        return f"{ip} ({hostname})"
    return ip


def _deduplicate_file(filepath: str) -> List[list]:
    """
    Parse a JSONL file, deduplicate by vulnerability name,
    and aggregate all affected hosts per unique vulnerability.
    Returns sorted rows ready for the spreadsheet.
    """
    # Key: vuln_name -> {severity, cvss, description, solution, hosts set}
    vuln_map: Dict[str, dict] = {}

    with open(filepath, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                continue

            vuln_name = str(rec.get("vulnerability_name", "") or "Unknown")
            host = _extract_host(rec)

            if vuln_name not in vuln_map:
                severity = str(rec.get("vulnerability_severity", "") or "N/A")
                try:
                    cvss3 = float(rec.get("vulnerability_cvss3_base_score", 0) or 0)
                except (ValueError, TypeError):
                    cvss3 = 0.0
                try:
                    cvss2 = float(rec.get("vulnerability_cvss2_base_score", 0) or 0)
                except (ValueError, TypeError):
                    cvss2 = 0.0
                cvss_score = cvss3 if cvss3 > 0 else cvss2

                description = str(rec.get("vulnerability_description", "") or "N/A")
                solution = str(rec.get("vulnerability_solution", "") or "N/A")

                vuln_map[vuln_name] = {
                    "severity": severity.capitalize(),
                    "cvss": cvss_score,
                    "description": description,
                    "solution": solution,
                    "hosts": set(),
                }

            vuln_map[vuln_name]["hosts"].add(host)

    # Convert to rows
    rows = []
    for vuln_name, data in vuln_map.items():
        hosts_sorted = sorted(data["hosts"])
        hosts_str = "\n".join(hosts_sorted)
        row = [
            vuln_name,
            data["severity"],
            data["cvss"],
            len(hosts_sorted),
            hosts_str,
            data["description"],
            data["solution"],
        ]
        rows.append(row)

    # Sort by criticality then CVSS desc then host count desc
    severity_order = {"Critical": 0, "High": 1, "Medium": 2}
    rows.sort(key=lambda r: (severity_order.get(r[1], 99), -(r[2] or 0), -r[3]))
    return rows


# ──────────────────────────────────────────────
# EXCEL GENERATION (write_only mode for performance with 800k rows)
# ──────────────────────────────────────────────

def _sanitize_sheet_title(name: str) -> str:
    """Remove characters invalid in Excel sheet names and truncate to 31 chars."""
    clean = re.sub(r'[/\\*?\[\]:]', '-', name)
    return clean[:31]


def _make_header_row(ws) -> list:
    """Create a styled header row."""
    cells = []
    for col_name in COLUMNS:
        cell = WriteOnlyCell(ws, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        cells.append(cell)
    return cells


def _make_data_row(ws, row_data: list) -> list:
    """Create a styled data row."""
    cells = []
    for col_idx, value in enumerate(row_data):
        cell = WriteOnlyCell(ws, value=value)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT
        cell.border = THIN_BORDER

        # Criticality is column index 1 now
        if col_idx == 1:
            severity_key = (str(value) or "").lower()
            if severity_key in SEVERITY_FILLS:
                cell.fill = SEVERITY_FILLS[severity_key]
                cell.font = SEVERITY_FONTS[severity_key]

        cells.append(cell)
    return cells


def _write_summary_sheet(wb: Workbook, group_stats: Dict[str, Dict[str, int]]):
    """Write the summary tab."""
    ws = wb.create_sheet(title="Summary")

    # Title
    title_cell = WriteOnlyCell(ws, value="Vulnerability Report by Asset Type")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="2F5496")
    ws.append([title_cell])

    # Date
    date_cell = WriteOnlyCell(ws, value=f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}")
    date_cell.font = Font(name="Calibri", size=11, italic=True)
    ws.append([date_cell])

    ws.append([])

    # Header
    summary_headers = ["Asset Group", "Total Vulnerabilities", "Critical", "High", "Medium"]
    header_cells = []
    for h in summary_headers:
        cell = WriteOnlyCell(ws, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        header_cells.append(cell)
    ws.append(header_cells)

    # Data rows
    sorted_groups = sorted(group_stats.keys(), key=lambda g: (g == "Other", g))
    total_all = total_crit = total_high = total_med = 0

    for group_name in sorted_groups:
        stats = group_stats[group_name]
        total = stats["total"]
        critical = stats["critical"]
        high = stats["high"]
        medium = stats["medium"]

        row_cells = []

        c = WriteOnlyCell(ws, value=group_name)
        c.font = DATA_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")
        row_cells.append(c)

        c = WriteOnlyCell(ws, value=total)
        c.font = DATA_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        row_cells.append(c)

        c = WriteOnlyCell(ws, value=critical)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = SEVERITY_FILLS["critical"] if critical > 0 else PatternFill()
        c.font = SEVERITY_FONTS["critical"] if critical > 0 else DATA_FONT
        row_cells.append(c)

        c = WriteOnlyCell(ws, value=high)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = SEVERITY_FILLS["high"] if high > 0 else PatternFill()
        c.font = SEVERITY_FONTS["high"] if high > 0 else DATA_FONT
        row_cells.append(c)

        c = WriteOnlyCell(ws, value=medium)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = SEVERITY_FILLS["medium"] if medium > 0 else PatternFill()
        c.font = SEVERITY_FONTS["medium"] if medium > 0 else DATA_FONT
        row_cells.append(c)

        ws.append(row_cells)
        total_all += total
        total_crit += critical
        total_high += high
        total_med += medium

    # Blank + totals
    ws.append([])
    totals_cells = []
    for val in ["TOTAL", total_all, total_crit, total_high, total_med]:
        c = WriteOnlyCell(ws, value=val)
        c.font = Font(name="Calibri", bold=True, size=11)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        totals_cells.append(c)
    totals_cells[0].alignment = Alignment(horizontal="left", vertical="center")
    ws.append(totals_cells)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12


def _write_data_sheet(wb: Workbook, sheet_title: str, rows: List[list]):
    """Write a single data tab."""
    ws = wb.create_sheet(title=sheet_title)

    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    ws.append(_make_header_row(ws))

    for i, row_data in enumerate(rows, start=1):
        ws.append(_make_data_row(ws, row_data))
        if i % 50000 == 0:
            print(f"    ... {i}/{len(rows)} rows written", flush=True)


def generate_report(group_files: Dict[str, str]) -> str:
    """Build the Excel workbook from temp JSONL files with deduplicated vulns."""
    wb = Workbook(write_only=True)

    # Deduplicate and compute stats per group
    print("Deduplicating vulnerabilities and computing stats...", flush=True)
    group_rows: Dict[str, List[list]] = {}
    group_stats: Dict[str, Dict[str, int]] = {}

    sorted_groups = sorted(group_files.keys(), key=lambda g: (g == "Other", g))
    for group_name in sorted_groups:
        filepath = group_files[group_name]
        rows = _deduplicate_file(filepath)
        group_rows[group_name] = rows

        critical = sum(1 for r in rows if r[1] == "Critical")
        high = sum(1 for r in rows if r[1] == "High")
        medium = sum(1 for r in rows if r[1] == "Medium")
        group_stats[group_name] = {
            "total": len(rows), "critical": critical, "high": high, "medium": medium
        }
        print(f"  {group_name}: {len(rows)} unique vulnerabilities", flush=True)

    # Summary sheet
    _write_summary_sheet(wb, group_stats)

    # Data sheets
    for group_name in sorted_groups:
        rows = group_rows[group_name]
        if not rows:
            continue
        print(f"  Writing sheet: {group_name} ({len(rows)} rows)...", flush=True)
        sheet_title = _sanitize_sheet_title(group_name)
        _write_data_sheet(wb, sheet_title, rows)

    # Save
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"runzero_vulnerability_report_{today}.xlsx"
    output_path = os.path.join(os.getcwd(), filename)
    print(f"Saving workbook to {output_path}...", flush=True)
    wb.save(output_path)
    return output_path


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

def main():
    if RUNZERO_ORG_TOKEN == "PASTE_YOUR_ORG_API_KEY_HERE":
        print("ERROR: Please paste your runZero Organisation API key into the script.")
        print("       Edit the RUNZERO_ORG_TOKEN variable near the top of this file.")
        sys.exit(1)

    tmp_dir = tempfile.mkdtemp(prefix="runzero_vuln_")

    try:
        group_files = fetch_vulnerabilities(tmp_dir)
    except requests.HTTPError as e:
        print(f"ERROR: API request failed — {e}")
        sys.exit(1)
    except requests.ConnectionError as e:
        print(f"ERROR: Could not connect to the runZero API — {e}")
        sys.exit(1)

    if not group_files:
        print("No medium, high, or critical vulnerabilities found. No report generated.")
        sys.exit(0)

    try:
        filepath = generate_report(group_files)
        print(f"\nReport saved: {filepath}")
    finally:
        # Cleanup temp files
        for fp in group_files.values():
            if os.path.exists(fp):
                os.remove(fp)
        try:
            os.rmdir(tmp_dir)
        except OSError:
            pass


if __name__ == "__main__":
    main()
